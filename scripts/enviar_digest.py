"""
Script de news semanal — Agência FOLKS.
Executado pelo GitHub Actions toda segunda-feira às 08h BRT.
Compila notícias da semana, gera PDF + XLSX e envia por email.
"""

import base64
import json
import os
import smtplib
import sys
import time
from datetime import datetime, timedelta
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from io import BytesIO  # usado pelo openpyxl

import pytz
from fpdf import FPDF, XPos, YPos
from jinja2 import Environment, FileSystemLoader
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(__file__))
from fontes import CORES_CATEGORIA, ORDEM_CATEGORIAS
from utils import BRT, NOTICIAS_DIR, log_erro, log_info

DIGESTS_DIR = os.path.join(os.path.dirname(__file__), "..", "digests")
TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), "..", "templates")
LOGO_PATH = os.path.join(os.path.dirname(__file__), "..", "logo-folks.png")

GMAIL_REMETENTE = os.environ.get("GMAIL_REMETENTE", "")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")
DESTINATARIOS_STR = os.environ.get("DESTINATARIOS", "")


MAX_POR_CATEGORIA = 15


def _carregar_noticias_semana() -> list:
    """Lê os JSONs dos últimos 7 dias e retorna lista mesclada sem duplicatas."""
    hoje = datetime.now(BRT).date()
    todas = {}
    for i in range(7):
        data = (hoje - timedelta(days=i)).strftime("%Y-%m-%d")
        caminho = os.path.join(NOTICIAS_DIR, f"{data}.json")
        if not os.path.exists(caminho):
            continue
        try:
            with open(caminho, encoding="utf-8") as f:
                noticias = json.load(f)
            for n in noticias:
                if n.get("id") and n["id"] not in todas:
                    todas[n["id"]] = n
        except Exception as e:
            log_erro(f"Erro ao ler {caminho}", e)
    return list(todas.values())


def _agrupar_por_categoria(noticias: list) -> dict:
    """
    Agrupa notícias por categoria, ordena cada grupo por prioridade e data,
    e limita a MAX_POR_CATEGORIA itens.
    """
    ordem_prio = {"Alto": 0, "Médio": 1, "Baixo": 2}
    grupos: dict = {cat: [] for cat in ORDEM_CATEGORIAS}

    for n in noticias:
        cat = n.get("categoria", "")
        if cat in grupos:
            grupos[cat].append(n)

    for cat in grupos:
        grupos[cat].sort(
            key=lambda x: (
                ordem_prio.get(x.get("prioridade", "Baixo"), 2),
                x.get("data_publicacao", ""),
            )
        )
        grupos[cat] = grupos[cat][:MAX_POR_CATEGORIA]

    return grupos


def _gerar_html(grupos: dict, semana_inicio: str, semana_fim: str, total: int) -> str:
    logo_b64 = ""
    if os.path.exists(LOGO_PATH):
        with open(LOGO_PATH, "rb") as f:
            logo_b64 = base64.b64encode(f.read()).decode("ascii")

    env = Environment(loader=FileSystemLoader(TEMPLATES_DIR))
    template = env.get_template("email_digest.html")
    return template.render(
        grupos=grupos,
        ordem_categorias=ORDEM_CATEGORIAS,
        cores_categoria=CORES_CATEGORIA,
        semana_inicio=semana_inicio,
        semana_fim=semana_fim,
        total_noticias=total,
        logo_b64=logo_b64,
    )


def _pdf_txt(texto: str) -> str:
    """Substitui caracteres fora do latin-1 por equivalentes ASCII para o PDF."""
    substituicoes = {
        "\u2014": "-",   # em dash —
        "\u2013": "-",   # en dash –
        "\u2019": "'",   # aspas tipograficas '
        "\u2018": "'",
        "\u201c": '"',   # aspas duplas tipograficas "
        "\u201d": '"',
        "\u2026": "...", # reticencias …
        "\u00e9": "e",   # fallback para outros fora do latin-1
    }
    for orig, sub in substituicoes.items():
        texto = texto.replace(orig, sub)
    # Remove qualquer char ainda fora do latin-1
    return texto.encode("latin-1", errors="replace").decode("latin-1")


def _gerar_pdf(grupos: dict, semana_inicio: str, semana_fim: str, total: int) -> bytes:
    """Gera PDF do digest usando fpdf2 (puro Python, sem dependências de sistema)."""
    icone_prio = {"Alto": "[ALTO]", "Médio": "[MEDIO]", "Baixo": "[BAIXO]"}
    cor_prio = {
        "Alto":  (229, 57, 53),
        "Médio": (249, 168, 37),
        "Baixo": (158, 158, 158),
    }

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_margins(15, 15, 15)
    pdf.add_page()

    NL = {"new_x": XPos.LMARGIN, "new_y": YPos.NEXT}

    # Cabeçalho — fundo vermelho FOLKS #A9170A = RGB(169, 23, 10)
    pdf.set_fill_color(169, 23, 10)
    pdf.rect(0, 0, 210, 44, "F")

    # Logo
    if os.path.exists(LOGO_PATH):
        pdf.image(LOGO_PATH, x=85, y=6, h=14)
        pdf.set_xy(0, 22)
    else:
        pdf.set_xy(0, 10)

    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(255, 252, 233)  # #FFFCE9
    pdf.cell(210, 8, "Agencia FOLKS - News Semanal", align="C", **NL)
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(210, 6, f"Contabilidade - Tributario - Previdenciario | {semana_inicio} a {semana_fim}", align="C", **NL)
    pdf.cell(210, 6, f"Total: {total} noticias coletadas", align="C", **NL)
    pdf.ln(12)

    pdf.set_text_color(0, 0, 0)

    for cat in ORDEM_CATEGORIAS:
        itens = grupos.get(cat, [])
        if not itens:
            continue

        # Título da categoria
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_fill_color(240, 240, 245)
        pdf.cell(0, 9, _pdf_txt(f"  {cat}  ({len(itens)} noticias)"), fill=True, **NL)
        pdf.ln(3)

        for n in itens:
            prio = n.get("prioridade", "Baixo")
            r, g, b = cor_prio.get(prio, (158, 158, 158))

            # Badge de prioridade
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_text_color(r, g, b)
            pdf.cell(0, 5, icone_prio.get(prio, ""), **NL)

            # Título
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(26, 26, 46)
            pdf.multi_cell(0, 5, _pdf_txt(n.get("titulo", "")), **NL)

            # Fonte e data
            pdf.set_font("Helvetica", "", 8)
            pdf.set_text_color(120, 120, 120)
            pdf.cell(0, 5, _pdf_txt(f"{n.get('fonte', '')}  |  {n.get('data_publicacao', '')[:10]}"), **NL)

            # Resumo
            resumo = n.get("resumo", "")[:280]
            if resumo:
                pdf.set_font("Helvetica", "", 9)
                pdf.set_text_color(70, 70, 70)
                pdf.multi_cell(0, 5, _pdf_txt(resumo), **NL)

            # Sugestão de pauta
            pauta = n.get("sugestao_pauta", "")
            if pauta:
                pdf.set_font("Helvetica", "I", 9)
                pdf.set_text_color(80, 80, 140)
                pdf.multi_cell(0, 5, _pdf_txt(f"Pauta: {pauta}"), **NL)

            # URL
            url = n.get("url", "")
            if url:
                pdf.set_font("Helvetica", "U", 8)
                pdf.set_text_color(0, 80, 200)
                pdf.cell(0, 5, url[:90], link=url, **NL)

            pdf.set_draw_color(220, 220, 220)
            pdf.line(15, pdf.get_y() + 2, 195, pdf.get_y() + 2)
            pdf.ln(6)
            pdf.set_text_color(0, 0, 0)

    return bytes(pdf.output())


def _gerar_xlsx(noticias: list, semana_inicio: str, semana_fim: str) -> bytes:
    """
    Gera plano editorial em Excel seguindo o padrão da skill /plano-editorial.
    Abas: Dashboard | Pauta da Semana | Referências
    """
    from datetime import datetime as _dt

    # --- Paleta FOLKS ---
    DARK_NAVY    = "A9170A"   # vermelho FOLKS
    GOLD         = "FFFCE9"   # bege FOLKS (texto nos cabeçalhos)
    LIGHT_GOLD   = "FAEEE6"   # bege rosado claro para KPIs
    LIGHT_GRAY   = "F2F2F2"
    WHITE        = "FFFFFF"

    COR_PILAR = {
        "Mercado":     "EDE1F5",
        "Educação":    "D9E8F5",
        "Autoridade":  "F5E9C8",
        "Dica Rápida": "FDE9CE",
    }
    COR_FORMATO = {"Feed": "D6E4F0", "Reels": "FDEDEC", "Stories": "E8F8F5"}
    COR_PRIO    = {"Alto": "FFCCCC", "Médio": "FFF9C4", "Baixo": "F5F5F5"}

    def _pilar(cat):
        return {"Tributário": "Mercado", "Legislação": "Mercado",
                "Contabilidade": "Educação", "Jurídico": "Autoridade"}.get(cat, "Mercado")

    def _formato(prio):
        return {"Alto": "Reels", "Médio": "Feed", "Baixo": "Stories"}.get(prio, "Feed")

    def _cta(prio):
        return {
            "Alto":  "Fale com um especialista sobre o impacto para o seu negócio",
            "Médio": "Salve esse post — você vai precisar dessa informação",
            "Baixo": "Compartilhe com quem precisa saber disso",
        }.get(prio, "Acompanhe nosso perfil para mais atualizações")

    def _hashtags(cat):
        return {
            "Tributário":    "#tributário #impostos #planejamentofiscal",
            "Legislação":    "#legislação #novasleis #direitotributário",
            "Contabilidade": "#contabilidade #contador #gestãocontábil",
            "Jurídico":      "#direitotributário #direitoprevidenciário #advocacia",
        }.get(cat, "#contabilidade #tributário")

    def _visual(cat, prio):
        if prio == "Alto":
            return "Fundo vermelho/laranja, ícone de alerta, texto em branco"
        return {
            "Tributário":    "Ícone de imposto/moeda, fundo azul escuro, texto objetivo",
            "Legislação":    "Ícone de documento/lei, fundo verde, destaque para o nome da lei",
            "Contabilidade": "Ícone de gráfico/calculadora, fundo azul claro, tom técnico",
            "Jurídico":      "Ícone de balança/justiça, fundo roxo, tom institucional",
        }.get(cat, "Arte clean com cores da marca")

    def _fill(color):
        return PatternFill("solid", fgColor=color)

    def _banner(ws, title, subtitle, n_cols):
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 20
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        c = ws.cell(row=1, column=1, value=title)
        c.font = Font(name="Arial", bold=True, size=16, color=GOLD)
        c.fill = _fill(DARK_NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        c2 = ws.cell(row=2, column=1, value=subtitle)
        c2.font = Font(name="Arial", size=11, color=WHITE)
        c2.fill = _fill(DARK_NAVY)
        c2.alignment = Alignment(horizontal="center", vertical="center")

    def _hdr(ws, row, col, value, bg=DARK_NAVY, fg=GOLD, size=9):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(name="Arial", bold=True, size=size, color=fg)
        cell.fill = _fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        return cell

    ordem_prio = {"Alto": 0, "Médio": 1, "Baixo": 2}
    noticias_ordenadas = sorted(
        noticias,
        key=lambda x: (ordem_prio.get(x.get("prioridade", "Baixo"), 2),
                       x.get("data_publicacao", "")),
    )

    wb = Workbook()

    # ── Dashboard ────────────────────────────────────────────────────────────
    ws_d = wb.active
    ws_d.title = "Dashboard"

    total    = len(noticias_ordenadas)
    por_cat  = {c: 0 for c in ORDEM_CATEGORIAS}
    por_prio = {"Alto": 0, "Médio": 0, "Baixo": 0}
    for n in noticias_ordenadas:
        por_cat[n.get("categoria", "Tributário")] = por_cat.get(n.get("categoria", "Tributário"), 0) + 1
        por_prio[n.get("prioridade", "Baixo")] = por_prio.get(n.get("prioridade", "Baixo"), 0) + 1

    _banner(ws_d, "Agência FOLKS — Plano Editorial de Notícias",
            f"Semana de {semana_inicio} a {semana_fim}  ·  {total} notícias mapeadas", 6)

    # KPIs linha 4-5
    kpis = [("Total", total), ("🔴 Alta", por_prio["Alto"]),
            ("🟡 Média", por_prio["Médio"]), ("⚪ Baixa", por_prio["Baixo"])]
    for col, (label, valor) in enumerate(kpis, 1):
        ws_d.row_dimensions[4].height = 22
        ws_d.row_dimensions[5].height = 36
        cl = ws_d.cell(row=4, column=col, value=label)
        cl.font = Font(name="Arial", bold=True, size=9, color="1A1A1A")
        cl.fill = _fill(LIGHT_GOLD)
        cl.alignment = Alignment(horizontal="center", vertical="center")
        cv = ws_d.cell(row=5, column=col, value=valor)
        cv.font = Font(name="Arial", bold=True, size=20, color=DARK_NAVY)
        cv.fill = _fill(WHITE)
        cv.alignment = Alignment(horizontal="center", vertical="center")

    # Notícias por categoria (linha 7+)
    ws_d.cell(row=7, column=1, value="NOTÍCIAS POR CATEGORIA").font = Font(
        name="Arial", bold=True, size=10, color=DARK_NAVY)
    for i, cat in enumerate(ORDEM_CATEGORIAS):
        r = 8 + i
        ws_d.row_dimensions[r].height = 18
        cc = ws_d.cell(row=r, column=1, value=cat)
        cc.font = Font(name="Arial", bold=True, size=10, color=DARK_NAVY)
        cc.fill = _fill(LIGHT_GRAY)
        cq = ws_d.cell(row=r, column=2, value=por_cat.get(cat, 0))
        cq.font = Font(name="Arial", size=10)
        cq.fill = _fill(WHITE)

    # Pilares (coluna 4-5)
    ws_d.cell(row=7, column=4, value="PILARES").font = Font(
        name="Arial", bold=True, size=10, color=DARK_NAVY)
    for i, (pilar, cor, desc) in enumerate([
        ("Mercado 📊",    "EDE1F5", "Tributário & Legislação"),
        ("Educação 🎓",   "D9E8F5", "Contabilidade"),
        ("Autoridade 💼", "F5E9C8", "Jurídico"),
    ]):
        r = 8 + i
        cp = ws_d.cell(row=r, column=4, value=pilar)
        cp.font = Font(name="Arial", bold=True, size=10)
        cp.fill = _fill(cor)
        cd = ws_d.cell(row=r, column=5, value=desc)
        cd.font = Font(name="Arial", size=9)
        cd.fill = _fill(WHITE)

    for col, w in [(1, 22), (2, 10), (3, 4), (4, 22), (5, 22), (6, 10)]:
        ws_d.column_dimensions[get_column_letter(col)].width = w

    # ── Pauta da Semana ──────────────────────────────────────────────────────
    ws_p = wb.create_sheet("Pauta da Semana")

    COLS   = ["Semana","Data","Dia","Formato","Pilar","Tema","Gancho",
              "Desenvolvimento","CTA","Hashtags","Sugestão Visual","Status","Fonte","URL"]
    WIDTHS = [14, 12, 11, 11, 16, 35, 45, 50, 35, 28, 32, 13, 20, 40]

    _banner(ws_p, "Agência FOLKS — Pauta da Semana",
            f"{semana_inicio} a {semana_fim}", len(COLS))

    ws_p.row_dimensions[3].height = 32
    for ci, name in enumerate(COLS, 1):
        _hdr(ws_p, 3, ci, name)
    ws_p.freeze_panes = "A4"
    ws_p.auto_filter.ref = f"A3:{get_column_letter(len(COLS))}3"

    for ri, n in enumerate(noticias_ordenadas, 4):
        ws_p.row_dimensions[ri].height = 52
        cat   = n.get("categoria", "Tributário")
        prio  = n.get("prioridade", "Baixo")
        pilar = _pilar(cat)
        fmt   = _formato(prio)

        try:
            d = _dt.fromisoformat(n.get("data_publicacao", "")[:10])
            data_s   = d.strftime("%d/%m/%Y")
            dia_s    = ["Segunda","Terça","Quarta","Quinta","Sexta","Sábado","Domingo"][d.weekday()]
            semana_s = f"Sem.{d.isocalendar()[1]} ({d.strftime('%d/%m')})"
        except Exception:
            data_s = n.get("data_publicacao", "")[:10]
            dia_s = ""
            semana_s = f"Semana de {semana_inicio}"

        valores = [semana_s, data_s, dia_s, fmt, pilar,
                   n.get("titulo", ""), n.get("sugestao_pauta", ""),
                   n.get("resumo", "")[:400], _cta(prio), _hashtags(cat),
                   _visual(cat, prio), prio, n.get("fonte", ""), n.get("url", "")]

        bg_row = LIGHT_GRAY if ri % 2 == 0 else WHITE
        for ci, valor in enumerate(valores, 1):
            cell = ws_p.cell(row=ri, column=ci, value=valor)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            col_name = COLS[ci - 1]
            if col_name == "Pilar":
                cell.fill = _fill(COR_PILAR.get(pilar, WHITE))
                cell.font = Font(name="Arial", bold=True, size=9)
            elif col_name == "Formato":
                cell.fill = _fill(COR_FORMATO.get(fmt, WHITE))
            elif col_name == "Status":
                cell.fill = _fill(COR_PRIO.get(prio, WHITE))
                cell.font = Font(name="Arial", bold=True, size=9)
            elif col_name == "Gancho":
                cell.fill = _fill(LIGHT_GOLD)
                cell.font = Font(name="Arial", bold=True, size=9, color=DARK_NAVY)
            else:
                cell.fill = _fill(bg_row)

    for ci, w in enumerate(WIDTHS, 1):
        ws_p.column_dimensions[get_column_letter(ci)].width = w

    # ── Referências ──────────────────────────────────────────────────────────
    ws_r = wb.create_sheet("Referências")
    _banner(ws_r, "Agência FOLKS — Referências", "Hashtags por categoria", 3)

    _hdr(ws_r, 4, 1, "CATEGORIA")
    _hdr(ws_r, 4, 2, "HASHTAGS SUGERIDAS")
    for i, (cat, tags) in enumerate([
        ("Tributário",    "#tributário #impostos #planejamentofiscal #reformatributária #direitofiscal"),
        ("Legislação",    "#legislação #novasleis #direitotributário #sefaz #prefeituradfortaleza"),
        ("Contabilidade", "#contabilidade #contador #gestãocontábil #escritóriocontábil #cfc"),
        ("Jurídico",      "#direitotributário #direitoprevidenciário #advocaciatributária #INSS"),
    ], 5):
        ws_r.row_dimensions[i].height = 18
        cc = ws_r.cell(row=i, column=1, value=cat)
        cc.font = Font(name="Arial", bold=True, size=10)
        cc.fill = _fill(LIGHT_GRAY)
        ct = ws_r.cell(row=i, column=2, value=tags)
        ct.font = Font(name="Arial", size=9)
        ct.fill = _fill(WHITE)
    for col, w in [(1, 20), (2, 70), (3, 10)]:
        ws_r.column_dimensions[get_column_letter(col)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _enviar_email(
    html_body: str,
    pdf_bytes: bytes,
    xlsx_bytes: bytes,
    assunto: str,
    destinatarios: list,
):
    """Envia o email com attachments via Gmail SMTP. Tenta até 3 vezes."""
    msg = MIMEMultipart("mixed")
    msg["From"] = GMAIL_REMETENTE
    msg["To"] = ", ".join(destinatarios)
    msg["Subject"] = assunto

    # Parte alternativa (texto simples + HTML)
    alt = MIMEMultipart("alternative")
    texto_simples = "News semanal disponível. Abra este email em um cliente que suporte HTML."
    alt.attach(MIMEText(texto_simples, "plain", "utf-8"))
    alt.attach(MIMEText(html_body, "html", "utf-8"))
    msg.attach(alt)

    # Anexo PDF
    pdf_part = MIMEApplication(pdf_bytes, _subtype="pdf")
    pdf_part.add_header("Content-Disposition", "attachment", filename="news_semanal.pdf")
    msg.attach(pdf_part)

    # Anexo XLSX
    xlsx_part = MIMEApplication(
        xlsx_bytes,
        _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    xlsx_part.add_header("Content-Disposition", "attachment", filename="pauta_semanal.xlsx")
    msg.attach(xlsx_part)

    for tentativa in range(1, 4):
        try:
            with smtplib.SMTP("smtp.gmail.com", 587, timeout=30) as smtp:
                smtp.ehlo()
                smtp.starttls()
                smtp.login(GMAIL_REMETENTE, GMAIL_APP_PASSWORD)
                smtp.sendmail(GMAIL_REMETENTE, destinatarios, msg.as_string())
            log_info(f"Email enviado com sucesso para {destinatarios}")
            return
        except Exception as e:
            log_erro(f"Tentativa {tentativa}/3 falhou", e)
            if tentativa < 3:
                time.sleep(5)

    raise RuntimeError("Falha ao enviar email após 3 tentativas.")


def main():
    log_info("Iniciando compilação da news semanal — Agência FOLKS")

    hoje = datetime.now(BRT).date()
    semana_inicio = (hoje - timedelta(days=6)).strftime("%d/%m/%Y")
    semana_fim = hoje.strftime("%d/%m/%Y")
    data_arquivo = hoje.strftime("%Y-%m-%d")

    noticias = _carregar_noticias_semana()
    log_info(f"Total de notícias carregadas: {len(noticias)}")

    if not noticias:
        log_info("Nenhuma notícia encontrada para a semana. News não enviada.")
        return

    grupos = _agrupar_por_categoria(noticias)
    total = sum(len(v) for v in grupos.values())

    # Gera artefatos
    log_info("Gerando HTML...")
    html = _gerar_html(grupos, semana_inicio, semana_fim, total)

    log_info("Gerando PDF...")
    pdf_bytes = _gerar_pdf(grupos, semana_inicio, semana_fim, total)

    log_info("Gerando planilha XLSX...")
    xlsx_bytes = _gerar_xlsx(noticias, semana_inicio, semana_fim)

    # Salva arquivos localmente (commitados no repo como arquivo)
    os.makedirs(DIGESTS_DIR, exist_ok=True)
    with open(os.path.join(DIGESTS_DIR, f"{data_arquivo}_news.html"), "w", encoding="utf-8") as f:
        f.write(html)
    with open(os.path.join(DIGESTS_DIR, f"{data_arquivo}_news.pdf"), "wb") as f:
        f.write(pdf_bytes)
    with open(os.path.join(DIGESTS_DIR, f"{data_arquivo}_pauta.xlsx"), "wb") as f:
        f.write(xlsx_bytes)
    log_info(f"Arquivos salvos em {DIGESTS_DIR}")

    # Envia email
    if not GMAIL_REMETENTE or not GMAIL_APP_PASSWORD or not DESTINATARIOS_STR:
        log_info("Credenciais de email não configuradas — email não enviado.")
        return

    destinatarios = [e.strip() for e in DESTINATARIOS_STR.split(",") if e.strip()]
    assunto = (
        f"[News Semanal] Notícias Contabilidade e Tributário "
        f"— {semana_inicio} a {semana_fim}"
    )
    _enviar_email(html, pdf_bytes, xlsx_bytes, assunto, destinatarios)
    log_info("News semanal concluída com sucesso.")


if __name__ == "__main__":
    main()
